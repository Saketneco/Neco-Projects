import requests
import os
from datetime import datetime
import logging
from openpyxl import Workbook
import openpyxl

# Configure logging
logging.basicConfig(level=logging.INFO)

def get_data(api_key, from_date, to_date):
    # Validate date range
    from_date_obj = datetime.strptime(from_date, '%Y-%m-%d')
    to_date_obj = datetime.strptime(to_date, '%Y-%m-%d')
    if (to_date_obj - from_date_obj).days > 7:
        logging.error("The difference between fromdate and todate cannot exceed 7 days.")
        return None

    # Construct the URL
    url = 'https://www.steelmint.com/api/public/v2019/pricesSMAPI'
    params = {
        'api_key': api_key,
        'fromdate': from_date,
        'todate': to_date
    }

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36'
    }

    try:
        response = requests.get(url, params=params, headers=headers)
        response.raise_for_status()
        data = response.json()
        if 'data' not in data:
            logging.warning("No data found in the response.")
            return None
        return data
    
    except requests.exceptions.HTTPError as err:
        logging.error(f"HTTP error occurred: {err}")
        logging.debug(f"Error Response: {response.text}")
        return None
    except requests.exceptions.RequestException as req_err:
        logging.error(f"Request error occurred: {req_err}")
        return None
    except Exception as err:
        logging.error(f"An unexpected error occurred: {err}")
        return None

def get_id_by_assessment(assessment_id, data_dict):
    assessment_id_str = str(assessment_id)
    if assessment_id_str in data_dict["AssessmentUniqueID"]:
        index = data_dict["AssessmentUniqueID"].index(assessment_id_str)
        return data_dict["ID"][index]
    else:
        logging.warning(f"Assessment ID '{assessment_id_str}' not found in data_dict.")
        return None

def save_to_xlsx(price_data, data_dict, directory):
    # Create the directory if it doesn't exist
    if not os.path.exists(directory):
        os.makedirs(directory)
    
    file_path = os.path.join(directory, 'bigmint_data.xlsx')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Price"

    # Write the header
    headers = ['PublishedDate', 'ID', 'Title', 'Currency', 'Price']
    sheet.append(headers)

    for entry in price_data.get('data', []):
        published_date = entry.get('PublishedDateTime', '').split(' ')[0].replace("-", "")  # Extract date part
        assessment_id = entry.get('ID')  # Ensure this is the right ID
        mapped_id = get_id_by_assessment(assessment_id, data_dict)  # Get mapped ID
        
        if mapped_id is None:
            logging.warning(f"No mapping found for Assessment ID '{assessment_id}'.")
        
        sheet.append([published_date, mapped_id, entry.get('Title'), entry.get('Currency'), entry.get('Price')])

    # Save the workbook
    workbook.save(file_path)
    logging.info(f"Data saved to '{file_path}'.")

    # Load the workbook to format specific columns
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Specify the columns you want to format as text
    text_columns = ["ID"]

    # Apply text formatting to specified columns, including the header
    for col in text_columns:
        col_index = headers.index(col) + 1  # Get column index (1-based)
        
        # Set the header to text format
        header_cell = ws.cell(row=1, column=col_index)  # Assuming headers are in the first row
        header_cell.number_format = '@'  # Set format to Text
        
        # Set the data cells to text format
        for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):  # Skip header
            for cell in row:
                cell.number_format = '@'  # Set format to Text
                if cell.value is not None:  # Avoid error if the cell is empty
                    cell.value = str(cell.value).zfill(18)  # Format to 18 characters with leading zeros

    # Save the updated workbook
    wb.save(file_path)
    logging.info(f"Data successfully written to '{file_path}'.")


# Example usage
if __name__ == "__main__":
    api_key = 'qkoeI6XB7YZhxupVE2Oe93'
    from_date = '2024-05-07'
    to_date = '2024-05-07'
    output_directory = 'D:\\USER PROFILE DATA\\Desktop\\Project-1\\Data\\old_data\\'  # Specify the directory where you want to save the file

    data_dict = {
        "AssessmentUniqueID": [
            "1254", "1251", "1255", "1660", "1091", "25", "1067", "1227",
            "1102", "1040", "1041", "1117", "1082", "69"
        ],
        "ID": [
            "000000050000121011", "000000050000121019", "000000050000121017",
            "000000050000121034", "IDRIPNA0003JNIL0000001", "000000050080001034",
            "000000050006102002", "RWRDPCS0320ROLL0005.50", "IPIGPST0001JNIL0000001",
            "000000050005101058", "000000050005101054", "000000050002357116",
            "IDRIPNA0001JNIL0000001", "IIOPPNA0002JNIL0000001"
        ]
    }

    # Fetch the price data
    price_data = get_data(api_key, from_date, to_date)

    # Print and save the output data
    if price_data:
        logging.info("Price Data fetched successfully:")
        logging.info(price_data)  # Print the entire response for inspection
        save_to_xlsx(price_data, data_dict, output_directory)
    else:
        logging.error("Failed to retrieve price data.")
