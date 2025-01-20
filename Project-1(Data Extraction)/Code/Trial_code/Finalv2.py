import os
import requests
import pdfplumber
from tabulate import tabulate
import re
import glob
import pandas as pd
import difflib
import logging
from datetime import datetime
import openpyxl
import http.client
from bs4 import BeautifulSoup
import Automation_pdf_extraction as Auto_pdf_ext


# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class PDFLinkProcessor:
    def __init__(self, base_url, files, download_directory):
        self.base_url = base_url
        self.files = files
        self.downloaded_files = []
        self.download_directory = download_directory
        os.makedirs(self.download_directory, exist_ok=True) 

    def download_pdf(self, url):
        """Download a PDF from the given URL."""
        try:
            file_name = os.path.basename(url)
            file_path = os.path.join(self.download_directory, file_name)
            response = requests.get(url)
            response.raise_for_status()
            
            with open(file_path, 'wb') as pdf_file:
                pdf_file.write(response.content)
            logging.info(f"Downloaded: {file_name}")
            return file_path
        
        except requests.exceptions.HTTPError as http_err:
          logging.error(f"HTTP error occurred while downloading {url}: {http_err}")
          self.handle_http_error(http_err, url)
        except requests.exceptions.RequestException as e:
          logging.error(f"Error downloading {url}: {e}")
          return None
      
    def handle_http_error(self, error, url):

        base_url = url.rsplit('/', 1)[0]  
        filename = url.rsplit('/', 1)[-1]  
        parts = filename[:-15].rsplit('-')  
        #print(parts)
    
        if len(parts) < 2:
            logging.error(f"Error processing URL: {url}")
            return
    
        parts[1] = parts[1].lower()  
        name_part = ''.join(parts)  
        full_name = name_part + filename[-15:]  
        modified_url = f"{base_url}/{full_name}"
    
        print(f"Handling HTTP error: {error} for URL: {url}")
        print(f"Trying modified URL: {modified_url}")


    def generate_pdf_urls_and_download(self):
        """Generate PDF URLs from the given input and download them."""
        for row in self.files:
            file_name = row[0].replace(" ", "-").title() + '-'
            file_date = row[1].replace("/", "-") + '.pdf'
            pdf_url = self.base_url + file_name + file_date
            #print(pdf_url)
            file_path = self.download_pdf(pdf_url)
            if file_path:
                self.downloaded_files.append(file_path)

    def extract_tables_from_pdf(self, pdf_path):
        """Extract tabular data from the PDF."""
        with pdfplumber.open(pdf_path) as pdf:
            all_tables = []
            for i, page in enumerate(pdf.pages):
                flat_table = self.extract_tables_from_page(page)
                if not flat_table:
                    logging.warning(f"No tables found on page {i + 1} of {pdf_path}.")
                    continue
                extracted_dates = self.modify_cells(flat_table)
                self.add_dates_currency_column(flat_table, extracted_dates)
                all_tables.extend(flat_table)

            self.save_to_xlsx(all_tables, pdf_path)

    def extract_tables_from_page(self, page):
        """Extract tables from a page."""
        tables = page.extract_tables()
        return [item for sublist in tables for item in sublist] if tables else []

    def modify_cells(self, flat_table):
        """Modify the cell based on a given condition."""
        extracted_dates = []
        base_pattern = "Basic Price Ex-Works"
        for row_index, row in enumerate(flat_table):
            for cell_index, cell in enumerate(row):
                if isinstance(cell, str) and difflib.SequenceMatcher(None, cell, base_pattern).ratio() > 0.6:
                    match = re.search(r"(\b\d{2}\.\d{2}\.\d{4}\b)", cell)
                    if match:
                        extracted_date = match.group(0)
                        extracted_dates.append(extracted_date)
                        #modified_string = re.sub(r"w\.e\.f\s*\b\d{2}\.\d{2}\.\d{4}\b", "", cell).strip()
                        flat_table[row_index][cell_index] = "Price"
        return extracted_dates

    def add_dates_currency_column(self, flat_table, extracted_dates):
        """Add the date column based on the extracted dates."""
        if extracted_dates:
            flat_table[0].insert(0,'PublishedDate')
            flat_table[0].insert(len(flat_table[0])-1,'Currency')
            for row in flat_table[1:]:
                if isinstance(row[0], str) and row[0].isdigit():
                    original_date = extracted_dates[0]
                    day, month, year = original_date.split('.')
                    reversed_date = f"{year}{month}{day}"
                    row.insert(0,reversed_date)
                    row.insert(len(row)-1,'INR')
                else:
                    row.append('')

    def save_to_xlsx(self, flat_table, pdf_path):
        """Save the extracted data in xlsx format."""
        df = pd.DataFrame(flat_table[1:], columns=flat_table[0])
        file_name = self._generate_xlsx_filename(pdf_path)
        xlsx_file = os.path.join(self.download_directory, f"{file_name}.xlsx")
        df.to_excel(xlsx_file, index=False)
        logging.info(f"Saved: {os.path.basename(xlsx_file)}")

    def _generate_xlsx_filename(self, pdf_path):
        """Generate a safe xlsx filename from the PDF path."""
        name = os.path.basename(pdf_path).rstrip('.pdf')
        words = [word for word in name.split('-') if not word.isdigit()]
        return '_'.join(words)

    def load_pdfs(self):
        """Load and process downloaded PDFs."""
        for pdf_file in self.downloaded_files:
            self.extract_tables_from_pdf(pdf_file)
                
    def extract_tables_from_page(self, page):
        """Extract tables from a page."""
        tables = page.extract_tables()
        return [item for sublist in tables for item in sublist] if tables else []    

    def update_dates_in_xlsx(self,directory):
        # Look for the filtered_data.csv file in the given directory
        for file_name in os.listdir(directory):
             print(directory)
             if file_name.endswith('.xlsx'):
                file_path = os.path.join(directory, file_name)
                print(file_path)
             # Check if the file exists
             if not os.path.isfile(file_path):
                print(f"File '{file_name}' not found in directory '{directory}'.")
                return
    
             # Load the CSV file
             df = pd.read_excel(file_path)
             # Check if the 'Dates' column exists
             current_date = datetime.now().strftime('%Y%m%d')
             
             
             if 'PublishedDate' in df.columns:
                 # Get the current date in YYYYMMDD format
                 # Update the 'Dates' column for all rows
                 df['PublishedDate'] = current_date
                 df.to_excel(file_path, index=False)
                 # Save the updated DataFrame back to the CSV file
                 
                 print("PublishedDate updated successfully.")
             else:
                 df.insert(0, 'PublishedDate', current_date)
                 df.to_excel(file_path, index=False)
                 print("Column 'PublishedDate' not found in the CSV file so created it.")
             wb = openpyxl.load_workbook(file_path)
             ws = wb.active

             # Specify the columns you want to format as text
             text_columns = ["ID"]
             
             # Apply text formatting to specified columns, including the header
             for col in text_columns:
                if col in df.columns:
                 col_index = df.columns.get_loc(col) + 1  # Get column index (1-based)
        
                # Set the header to text format
                 header_cell = ws.cell(row=1, column=col_index)  # Assuming headers are in the first row
                 header_cell.number_format = '@'  # Set format to Text

                # Set the data cells to text format
                 for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):  # Skip header
                    for cell in row:
                        cell.number_format = '@'  # Set format to Text
                        cell.value = f"{int(cell.value):0>18}"

                # Save the updated workbook
                wb.save(file_path)
                print(f"Data successfully written to {file_path}")




class CSVManager:
    @staticmethod
    def list_xlsx_files(directory):
        """List the xlsx files in the given directory."""
        return glob.glob(os.path.join(directory, '*.xlsx'))

    @staticmethod
    def read_xlsx_files(xlsx_files):
        """Read the xlsx files and return data matrices."""
        data_matrices = {}
        for filename in xlsx_files:
            df = pd.read_excel(filename)
            data_matrices[os.path.basename(filename)] = {
                'data': df.values.tolist(),
                'columns': df.columns.tolist()
            }
        return data_matrices

    @staticmethod
    def add_ID_column(data_matrices, add_material_code):
        """Add the 'Material Code' column to the data matrices."""
        for file_name, data in data_matrices.items():
            if file_name in add_material_code:
                for i in range(len(add_material_code[file_name])):
                    material_code_name = add_material_code[file_name][i][0]
                    material_code_value = add_material_code[file_name][i][1]

                    data['columns'].insert(1, material_code_name)
                    for row in data['data']:
                        row.insert(1, material_code_value)
        return data_matrices

    @staticmethod
    def filter_data(data, columns, conditions):
        """Filter the data based on the given conditions."""
        #print(data)
        #print(columns)
        df = pd.DataFrame(data, columns=columns)
        temp_list = []
        #print(df['Sl.no.'])
        for column_name, value in conditions:
            if column_name not in df.columns:
                logging.warning(f"Column '{column_name}' does not exist in the DataFrame.")
                return pd.DataFrame()
           # print(column_name)
           # print(value)
           # elif df[column]==value:
                
            df = df[df[column_name] == value].drop(columns=[column_name,'Sl.no.'])
           # temp_list = df.to_numpy().tolist()
          #  print(df)
        return df
    
    @staticmethod
    def remove_column(data,columns):
       sl_no_index = columns.index('Sl.no.')
       new_data = [row[:sl_no_index] + row[sl_no_index + 1:] for row in data]
       new_columns = [col for col in columns if col != 'Sl.no.']
       
       return new_data,new_columns
   
    @staticmethod
    def apply_filter_conditions(data_matrices, filter_conditions):
        """Apply filter conditions to the data matrices."""
        filtered_dfs = []
        for filename, conditions in filter_conditions.items():
            if filename in data_matrices:
                
                data = data_matrices[filename]['data']
                columns = data_matrices[filename]['columns']
                #new_data,new_column = CSVManager.remove_column(data,columns)
                filtered_df = CSVManager.filter_data(data, columns, conditions)
                #print(filtered_df)
                #new_data,new_column = CSVManager.remove_column(data,columns)
                #print(filtered_df)
                if len(filtered_df) != 0:
                    filtered_dfs.append(filtered_df)
        return filtered_dfs


    
    @staticmethod
    def concatenate_and_save(filtered_dfs, output_file):
        """Concatenate filtered dataframes and save to excel."""
        text_columns =["ID"]
        if filtered_dfs:
            final_filtered_df = pd.concat(filtered_dfs, ignore_index=True)
            if 'Sl.no.' in final_filtered_df.columns:
                final_filtered_df['Sl.no.'] = range(1, len(final_filtered_df) + 1)
            final_filtered_df.to_excel(output_file, index=False, engine='openpyxl')

            # Optionally, set the format to Text in the saved Excel file
            # wb = openpyxl.load_workbook(output_file)
            # ws = wb.active
    
            # # Apply text formatting to specified columns
            # for col in text_columns:
            #     if col in final_filtered_df.columns:
            #         col_index = final_filtered_df.columns.get_loc(col) + 1  # Get column index (1-based)
            #         for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):  # Skip header
            #             for cell in row:
            #                 cell.number_format = '@'  # Set format to Text

            # wb.save(output_file)
            # print(f"Data successfully written to {output_file}")
        else:
            logging.warning("No data matched the filtering conditions.")

    @staticmethod
    def delete_pdf_files(directory):
        """Delete all PDF files in the given directory."""
        if not os.path.exists(directory):
            logging.warning(f"The directory '{directory}' does not exist.")
            return

        for filename in os.listdir(directory):
            if filename.endswith('.pdf'):
                file_path = os.path.join(directory, filename)
                try:
                    os.remove(file_path)
                    logging.info(f"Deleted: {file_path}")
                except Exception as e:
                    logging.error(f"Error deleting {file_path}: {e}")

class PDFTableExtractorPlatts:
    def __init__(self, pdf_path,download_directory,filter_data,Platts_Id):
        self.pdf_path = pdf_path
        self.download_directory = download_directory
        self.filter_data = filter_data
        self.Platts_Id = Platts_Id

    def extract_table_platts_pdf(self, page_number):
        with pdfplumber.open(self.pdf_path) as pdf:
            page = pdf.pages[page_number - 1]  # Convert to 0-based index
            tables = self.extract_tables(page)

            print(f"Number of tables extracted: {len(tables)}\n")

            if tables:
                specific_table_index = 0
                if len(tables) > specific_table_index:
                    table = tables[specific_table_index]
                    output_list = self.process_table(table)
                    self.save_to_excel_with_text_conversion(output_list, "Platts_data.xlsx")
                else:
                    print(f"No table found at index {specific_table_index}.")
            else:
                print("No tables found on this page.")

    def extract_tables(self, page):
        return page.extract_tables(table_settings={
            "vertical_strategy": "text",
            "horizontal_strategy": "lines_strict",
            "explicit_vertical_lines": [350, 310, 270, 220, 170],
            "explicit_horizontal_lines": [113],
            "snap_tolerance": 8,
            "join_tolerance": 1,
            "edge_min_length": 2,
            "min_words_vertical": 2,
            "min_words_horizontal": 2,
            "intersection_tolerance": 2,
            "text_tolerance": 2,
        })

    def process_table(self, table):
        df = pd.DataFrame(table)
        table_as_list = df.values.tolist()
        output_list = {}

        for row in self.filter_data.keys():
            for i, row2 in enumerate(table_as_list):
                if row == row2[0]:
                    column_list = []
                    for j, column in enumerate(table_as_list[1]):
                        if self.filter_data[row][0] == column.replace("\n", " ") or \
                           self.filter_data[row][1] == column.replace("\n", " "):
                            column_list.append(table_as_list[i][j])
                            
                    output_list[row] = column_list    
                   # print(output_list)
                    break

        return output_list

    def insert_column(self,row_list,columns):
        row_list = [list(item) for item in row_list]
        for key,value in self.Platts_Id.items():
            if value[0][0] not in columns:
                columns.insert(0,value[0][0])
                columns.insert(len(columns)-2,'Currency')
                
            for row in row_list:
                if row[0]== key:
                    #print()
                    row.insert(0,value[0][1])
                    row.insert(len(row)-2,"USD")
                    break   
            
        return row_list , columns       
            
    # def save_to_csv(self, output_list, output_filename):
    #     columns = ["Comodity Name"] + list(self.filter_data.values())[0]
    #     row_list = [(name, *value) for name, value in output_list.items()]
    #     #print(row_list)
    #     row_list,columns = self.insert_column(row_list,columns)
    #     df = pd.DataFrame(row_list, columns=columns)
    #     #df = df.convert_dtypes(convert_string=True)
    #     #print(df.dtypes)
    #     output_path = f"{self.download_directory}{output_filename}"
    #     df.to_csv(output_path, index=False)
    #     #df.to_excel(output_path, index=False, engine='openpyxl')
    #     print(f"Data successfully written to {output_filename}")


    def save_to_excel_with_text_conversion(self, output_list, output_filename):
        # Define the columns based on provided data
        text_columns =["ID"]
        columns = ["Commodity Name"] + list(self.filter_data.values())[0]
        row_list = [(name, *value) for name, value in output_list.items()]

        # Insert additional columns if needed
        row_list, columns = self.insert_column(row_list, columns)

        # Create a DataFrame
        df = pd.DataFrame(row_list, columns=columns)

        # Ensure the output filename has the correct extension
        if not output_filename.endswith('.xlsx'):
            output_filename += '.xlsx'

        # Create the full output path
        output_path = f"{self.download_directory}{output_filename}"

        # Convert specified columns to text (if any)
        if text_columns:
            for col in text_columns:
                if col in df.columns:
                    df[col] = df[col].astype(str)  # Convert column to string

        # Save to Excel
        df.to_excel(output_path, index=False, engine='openpyxl')

        # # Optionally, set the format to Text in the saved Excel file
        # wb = openpyxl.load_workbook(output_path)
        # ws = wb.active
    
        # # Apply text formatting to specified columns
        # for col in text_columns:
        #     if col in df.columns:
        #         col_index = df.columns.get_loc(col) + 1  # Get column index (1-based)
        #         for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):  # Skip header
        #             for cell in row:
        #                 cell.number_format = '@'  # Set format to Text

        # wb.save(output_path)
        # print(f"Data successfully written to {output_path}")

class LMEDataFetcher:
    def __init__(self, metal,download_directory,LME_data):
        self.connection = http.client.HTTPSConnection('www.lme.com')
        self.headers = {
            'User-Agent': 'Mozilla/5.0',
            'Referer': 'https://www.lme.com/',
            'Accept-Language': 'en-US,en;q=0.9',
        }
        self.metal = metal
        self.download_directory = download_directory
        self.LME_data = LME_data

    def fetch_data(self):
        self.connection.request('GET', f'/Metals/Non-ferrous/LME-{self.metal}', headers=self.headers)
        response = self.connection.getresponse()
        return response.read() if response.status == 200 else None

    def parse_data(self, data):
        soup = BeautifulSoup(data, 'html.parser')
        price_element = soup.find('div', class_='hero-metal-data__data')
        if price_element:
            return (
                price_element.find('span', class_='hero-metal-data__number').text.strip()
            )
        return None, None

    def close(self):
        self.connection.close()
        
    def save_file(self,price):
        if price:
            self.LME_data["Price"] = price
            
        df = pd.DataFrame([self.LME_data])
        name = "LME_data.xlsx"
        output_path = f"{self.download_directory}{name}"
        df.to_excel(output_path, index=False)
        print("Data has been written to", os.path.basename(output_path))

class PDFDataExtractorArgus:
    def __init__(self, pdf_path,download_directory,Argus_data):
        self.pdf_path = pdf_path
        self.table_settings = {
            "text_x_tolerance": 2,
            "text_y_tolerance": 2,
            "explicit_vertical_lines": [306, 348, 367, 394, 489, 518, 550]
        }
        self.download_directory = download_directory
        self.Argus_data = Argus_data

    def extract_data(self):
        with pdfplumber.open(self.pdf_path) as pdf:
            # Select the first page (index 0)
            page = pdf.pages[0]
            
            # Extract the table with the specified settings
            tables = page.extract_tables(table_settings=self.table_settings)

            if tables:
                first_table = tables[0]
                columns = ['Energy', 'Basis', 'Timing', 'Port', '', 'Price', None, '', '±']

                name = None
                price = None

                for i, row in enumerate(first_table):
                    if row[0].lower() == 'south africa':
                        # Ensure i + 1 is within bounds
                        if i + 1 < len(first_table):
                            name = first_table[i + 1][columns.index('Port')]
                            self.Argus_data['Price'] = first_table[i + 1][columns.index('Price')]
                            price =self.Argus_data['Price']
                            print(price)
                            break  # Exit the loop once the values are found

                return name, price
            else:
                print("No tables found on the page.")
                return None, None

    def save_to_excel(self):
        # Create a DataFrame from the dictionary
        df = pd.DataFrame([self.Argus_data])
        name = "Argus_data.xlsx"
        output_path = f"{self.download_directory}{name}"
        # Write the DataFrame to an Excel file
        df.to_excel(output_path, index=False)
        print("Data has been written to", os.path.basename(output_path))

    


def run_data_filtering_process(directory="D:\\USER PROFILE DATA\\Desktop\\Project\\Data\\", filter_conditions=None):
    """Run the data filtering process."""
    if filter_conditions is None:
        filter_conditions = {
      'Ingot.xlsx': [('Product Code', 'IC20')],
      'Sowingot.xlsx': [('Product Code', 'SC20')],
      'Sow_Ingot.xlsx': [('Product Code', 'SC20')],
      'Wirerod.xlsx': [('Product Code', 'WF10')],
      'Wire_Rod.xlsx': [('Product Code', 'WF10')]
    }
    
    add_material_code = {
        'Ingot.xlsx': [("ID", "50000121043")],
        'Sowingot.xlsx': [("ID", "50000121116")],
        'Sow_Ingot.xlsx': [("ID", "50000121116")],
        'Wirerod.xlsx': [("ID", "50000121024")],
        'Wire_Rod.xlsx': [("ID", "50000121024")]
    }

    csv_files = CSVManager.list_xlsx_files(directory)
    data_matrices = CSVManager.read_xlsx_files(csv_files)
    CSVManager.delete_pdf_files(directory)

    updated_matrices = CSVManager.add_ID_column(data_matrices, add_material_code)
    
    filtered_dfs = CSVManager.apply_filter_conditions(updated_matrices, filter_conditions)
    
    output_file = os.path.join(directory, 'Nalco_data.xlsx')
    CSVManager.concatenate_and_save(filtered_dfs, output_file)

def main():
    
    current_date = datetime.now()
    formatted_date = current_date.strftime("%d/%m/%Y")
    #print(formatted_date)

    # Define custom data
    files = [
        ['sow ingot', formatted_date],
        ['ingot', formatted_date],
        ['wire rod', formatted_date]
    ]
    base_url = "https://d2ah634u9nypif.cloudfront.net/wp-content/uploads/2019/01/"
    pdf_path = "D:\\USER PROFILE DATA\\Desktop\\Project-1\\Data\\pdf\\"

    
    download_directory = "D:\\USER PROFILE DATA\\Desktop\\Project-1\\Data\\"
    
    filter_data = {
            "Premium Low Vol": ["FOB Australia", "CFR India"],
            "Low Vol HCC": ["FOB Australia", "CFR India"],
            "Low Vol PCI": ["FOB Australia", "CFR India"],
            "Mid Vol PCI": ["FOB Australia", "CFR India"],
            "Semi Soft": ["FOB Australia", "CFR India"],
        }
    
    Platts_Id = {
            "Premium Low Vol": [("ID","50000141001")],
            "Low Vol HCC": [("ID","50000141012")],
            "Low Vol PCI": [("ID","50000151001")],
            "Mid Vol PCI": [("ID","50000151002")],
            "Semi Soft": [("ID","50000151003")],
        }
    
    LME_data = {
        "PublishedDate": datetime.now().strftime("%Y%m%d"),
        "ID": "50000121042",
        "Commodity Name": "Nickel Plate",
        "Currency": "USD",
        "Price": ""
    }
    
    Argus_data = {
        "PublishedDate": datetime.now().strftime("%Y%m%d"),
        "ID": "50000111006",
        "Commodity Name": "Steam Coal - 6000 NAR",
        "Currency": "USD",
        "Price": ""
    }
    

    
    #files = [['sow ingot', '01/10/2024'], ['ingot', '01/10/2024'], ['wire rod', '01/10/2024']]
    
    logging.basicConfig(level=logging.INFO)
    pdf_processor = PDFLinkProcessor(base_url, files, download_directory)
    pdf_processor.generate_pdf_urls_and_download()
    pdf_processor.load_pdfs()
    run_data_filtering_process(download_directory)
    #pdf_processor.update_dates_in_xlsx(download_directory)
    
    
    
    
    #Extarct latest file from director for data extraction
    file_loader = Auto_pdf_ext.LatestFileLoader(pdf_path)
    latest_ict_file, latest_cdi_file = file_loader.load_latest_files()
    if latest_ict_file:
        print(f"The latest ICT file is: {latest_ict_file}")
    else:
        print("No matching ICT files found.")

    if latest_cdi_file:
        print(f"The latest CDI file is: {latest_cdi_file}")
    else:
        print("No matching CDI files found.")
        
        
        
    
    #Platts_data_extraction
    pdf_extractor = PDFTableExtractorPlatts(latest_ict_file,download_directory,filter_data,Platts_Id)
    pdf_extractor.extract_table_platts_pdf(7)
    # updated_matrices = CSVManager.add_ID_column(data_matrices, add_material_code)
   
   
   

    
    #LME_price_extraction
    fetcher = LMEDataFetcher('Nickel',download_directory,LME_data)
    data = fetcher.fetch_data()

    if data:
        price = fetcher.parse_data(data)
        fetcher.save_file(price)
    fetcher.close()
    
    
    
    
    
    #Argus_data_extraction
    extractor = PDFDataExtractorArgus(latest_cdi_file, download_directory,Argus_data)
    name, price = extractor.extract_data()
    
    if name and price is None :
        print("Not able to extract data from argus file")
        
    extractor.save_to_excel()




    # Update current date
    pdf_processor.update_dates_in_xlsx(download_directory)
    
if __name__ == "__main__":
    main()
