import os
import glob
import pandas as pd
import openpyxl
from datetime import datetime

class ExcelMereger_combined_history:
    def __init__(self, directory):
        """
        Initialize the class with the directory where the files are stored.
        - directory: The directory where Excel files are located.
        
        Example:
        directory = "G:\\Shared drives\\Metal Prices\\Market_price\\"
        merger = ExcelMereger(directory)
        """
        self.directory = directory
        self.pattern = os.path.join(directory, 'combined_data_*.xlsx')

    def combine_files(self):
        """
        Combine all Excel files that match the pattern defined in the directory into one Excel file. 
        The files are sorted by date extracted from their filenames before combining.
        
        Example:
        If the directory contains:
        - combined_data_01-12-2023.xlsx
        - combined_data_05-12-2023.xlsx
        - combined_data_10-12-2023.xlsx
        - combined_data_09-12-2023.xlsx
        - combined_data_03-12-2023.xlsx 
        
        After calling combine_files, the files will be sorted in this order:
        - combined_data_01-12-2023.xlsx
        - combined_data_03-12-2023.xlsx
        - combined_data_05-12-2023.xlsx
        - combined_data_09-12-2023.xlsx
        - combined_data_10-12-2023.xlsx 
        
        
        dataframe = 
        """
        
        files = glob.glob(self.pattern)
        

        files.sort(key=self.extract_date)
        file_path = os.path.join(self.directory, 'temp.xlsx')

        if os.path.exists(file_path):
            files.append(file_path)
        else:
            print("\033[91mNo file found\033[0m")
            
        for file in files:
            print(file)

        # List to hold DataFrames
        dataframes = []
        
        for filename in files:
            # Read each Excel file and append the DataFrame to the list
            df = pd.read_excel(filename)
            dataframes.append(df)
            
        
        
        if dataframes:
            # Concatenate all DataFrames
            combined_df = pd.concat(dataframes, ignore_index=True)

            # Name of the output file
            output_file = os.path.join(self.directory, 'Combined_history.xlsx')

            # Write the combined DataFrame to an Excel file
            combined_df.to_excel(output_file, index=False)
            #df.to_excel(output_file, index=False)

            wb = openpyxl.load_workbook(output_file)
            ws = wb.active

            # Specify the columns you want to format as text
            text_columns = ["ID"]

            # Apply text formatting to specified columns, including the header
            for col in text_columns:
                if col in df.columns:
                    col_index = df.columns.get_loc(col) + 1  # Get column index (1-based)

                    # Set the header to text format
                    header_cell = ws.cell(row=1, column=col_index)  # Assuming headers are in the first row
                    header_cell.number_format = '@'  # Set format to Text

                    # Set the data cells to text format
                    for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):  # Skip header
                        for cell in row:
                           cell.number_format = '@'  # Set format to Text
                           cell.value = f"{cell.value:0>18}"

            # Save the updated workbook
            wb.save(output_file)
            print(f"Data successfully written to {os.path.basename(output_file)}")
            print(f"Combined {len(files)} files into {output_file}")
        else:
            print("No files found to combine.")

    def extract_date(self, filename):
        # Extract the date part from the filename
        basename = os.path.basename(filename)
        # Split the filename to find the date string
        date_str = basename.split('_')[-1].replace('.xlsx', '')  # Get the date part
        return datetime.strptime(date_str, '%d-%m-%Y')  # Convert to a datetime object

# # Usage example
# if __name__ == "__main__":
#     directory = "G:\\Shared drives\\Metal Prices\\Market_price\\"  # Change this to your directory path
    
    
