import os
import pandas as pd
from datetime import datetime
import openpyxl
import re

class MarketPriceChecker:
    def __init__(self, directory, titles):
        self.directory = directory
        self.titles = titles
        
    def get_all_files(self):
        files = os.listdir(self.directory)
        pattern = r"Combined_data_\d{2}-\d{2}-\d{4}\.xlsx"  # Pattern for "Combined_data_dd-mm-yyyy.xlsx"
        return [f for f in files if re.match(pattern, f)]
    
    def get_latest_file(self):
        files = self.get_all_files()
        
        # Extract date from filenames and sort by date
        files_with_date = []
        for file in files:
            try:
                date_str = file.split('_')[2].replace('.xlsx', '')
                file_date = datetime.strptime(date_str, "%d-%m-%Y")
                files_with_date.append((file, file_date))
            except ValueError:
                continue
        
        # Sort by date in descending order
        sorted_files = sorted(files_with_date, key=lambda x: x[1], reverse=True)
        return sorted_files[0][0] if sorted_files else None
    

    
    def check_titles_in_file(self, file_path):
        file_path = os.path.join(self.directory, file_path)
        df = pd.read_excel(file_path)
        
        # Check for titles in the dataframe
        missing_titles = {title: None for title in self.titles if title not in df['Title'].values}
        
        return missing_titles, df
    
    def find_titles_in_previous_files(self, latest_file):
        files = self.get_all_files()
        
        # Extract date from filenames and sort by date
        files_with_date = []
        for file in files:
            try:
                date_str = file.split('_')[2].replace('.xlsx', '')
                file_date = datetime.strptime(date_str, "%d-%m-%Y")
                files_with_date.append((file, file_date))
            except ValueError:
                continue

        # Sort the files by date in descending order
        sorted_files = sorted(files_with_date, key=lambda x: x[1], reverse=True)

        # Find the index of the latest file
        latest_file_index = next((i for i, item in enumerate(sorted_files) if item[0] == latest_file), None)

        # Iterate over current file and previous files to find missing titles
        for i in range(latest_file_index, len(sorted_files)):
            current_file = sorted_files[i][0]
            print(f"Checking file: {current_file}")
            missing_titles, df = self.check_titles_in_file(current_file)
            
            if not missing_titles:
                print("All titles are present in the current file.")
                break  # Stop if all titles are found in the current file
            
            # Search for missing titles in previous files
            for title in missing_titles:
                print(f"\nTitle {title} missing in {current_file}. Checking previous files.")
      
                #print(f"Latest file index: {latest_file_index}")

                
                for j in range(i+1, len(sorted_files)):  # Iterate in reverse order through previous files
                    prev_file = sorted_files[j][0]
                   # print(f"Checking previous file: {prev_file}")
                    prev_missing_titles, prev_df = self.check_titles_in_file(prev_file)
                    
                    if title not in prev_missing_titles:  # Title found in previous file
                        print(f"Found {title} in {prev_file}. Copying data.\n")
                        prev_title_row = prev_df[prev_df['Title'] == title]
                        df = pd.concat([df, prev_title_row], ignore_index=True)
                        #print(prev_df)
                        break  # Stop after finding the title in a previous file
                    else:
                        print(f"Title {title} not found in {prev_file}.")
            
            
            output_file = os.path.join(self.directory, f"{current_file}")
            # Save the updated dataframe to a new file
            df.to_excel(output_file, index=False)
            
            wb = openpyxl.load_workbook(output_file)
            ws = wb.active

            # Specify the columns you want to format as text
            text_columns = ["ID"]

            # Apply text formatting to specified columns, including the header
            for col in text_columns:
                if col in df.columns:
                    col_index = df.columns.get_loc(col) + 1  # Get column index (1-based)

                    # Set the header to text format
                    header_cell = ws.cell(row=1, column=col_index)  # Assuming headers are in the first row
                    header_cell.number_format = '@'  # Set format to Text

                    # Set the data cells to text format
                    for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):  # Skip header
                        for cell in row:
                           cell.number_format = '@'  # Set format to Text
                           cell.value = f"{cell.value:0>18}"

            # Save the updated workbook
            wb.save(output_file)
            print(f"Data successfully written to {os.path.basename(output_file)}")
    
    def extract_titles_to_temp(self, final_file):
        
        # Get today's date in the format 'yyyy-mm-dd'
        today_date = datetime.today().strftime('%Y%m%d')
        
        if datetime.today().weekday() == 5 or datetime.today().weekday() == 6:
            print("Today is Saturday or Sunday, returning.")
            return
        else:
            # Continue with your code
            print(f"Today is {today_date}, continuing execution.")
        
        
        # Read the final file to get all titles
        df = pd.read_excel(final_file)

        # Filter the DataFrame to only include rows where 'Title' matches any of the titles in self.titles
        filtered_df = df[df['Title'].isin(self.titles)]

        

        # Add today's date to the 'PublishedDate' column for each row
        filtered_df['PublishedDate'] = today_date

        # Save the filtered rows to a new file 'temp.xlsx'
        temp_file_path = os.path.join(self.directory, "temp.xlsx")
        filtered_df.to_excel(temp_file_path, index=False)
        
        wb = openpyxl.load_workbook(temp_file_path)
        ws = wb.active

        # Specify the columns you want to format as text
        text_columns = ["ID"]

        # Apply text formatting to specified columns, including the header
        for col in text_columns:
            if col in df.columns:
                col_index = df.columns.get_loc(col) + 1  # Get column index (1-based)

                # Set the header to text format
                header_cell = ws.cell(row=1, column=col_index)  # Assuming headers are in the first row
                header_cell.number_format = '@'  # Set format to Text

                # Set the data cells to text format
                for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):  # Skip header
                    for cell in row:
                       cell.number_format = '@'  # Set format to Text
                       cell.value = f"{cell.value:0>18}"

        # Save the updated workbook
        wb.save(temp_file_path)

        print(f"Filtered rows successfully written to {temp_file_path}")
        
        
    def delete_temp_file(self):
        # Path to the temp.xlsx file in the given directory
        temp_file_path = os.path.join(self.directory, 'temp.xlsx')

        # Check if the file exists
        if os.path.exists(temp_file_path):
            # Delete the file
            os.remove(temp_file_path)
            print(f"File 'temp.xlsx' has been deleted from {self.directory}.")
        else:
            print(f"File 'temp.xlsx' does not exist in {self.directory}.")

    
    def execute(self):
        latest_file = self.get_latest_file()
        print(f"Checking for the latest file: {latest_file}")
        
        if latest_file:
            print(f"Latest file found: {latest_file}")
            self.find_titles_in_previous_files(latest_file)
            # After processing the files, extract titles and save them to a temp file
            final_file = os.path.join(self.directory, f"{latest_file}")

            self.extract_titles_to_temp(final_file)
        else:
            print("No valid files found in the directory.")

# # Define the directory where your files are stored
# directory = "D:\\USER PROFILE DATA\\Desktop\\Project-1(Data Extraction)\\Data\\Market_price\\"
# titles_to_check = [
#     "Premium Low Vol", "Low Vol HCC", "Low Vol PCI", "Mid Vol PCI",
#     "Semi Soft", "Richards Bay-India West", "Australia-India", "Metallurgical coke(64/62)"
# ]

# # Create the MarketPriceChecker instance and execute
# market_price_checker = MarketPriceChecker(directory, titles_to_check)
# market_price_checker.execute()
