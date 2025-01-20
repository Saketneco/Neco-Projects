import os
import re
import time  # Import time module for sleep functionality
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import openpyxl


class CurrencyExchangeScraper:
    def __init__(self, Exchange_rate, file_paths):
        """
        Initializes the scraper with currency data and the path to the Excel file.
        """
        self.Exchange_rate = Exchange_rate
        self.file_paths = file_paths

        # Current date and date for yesterday
        self.current_date = datetime.now()
        self.yesterday_date = self.current_date - timedelta(days=1)
        self.yesterday_date_str = self.yesterday_date.strftime('%d-%m-%Y')
        self.current_month_year = self.yesterday_date.strftime("%B %Y")

        # Configure Chrome options for the WebDriver
        self.chrome_options = self.configure_chrome_options()

        # Initialize the WebDriver
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=self.chrome_options)

    def configure_chrome_options(self):
        """
        Configures and returns Chrome options for the WebDriver.
        """
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--enable-unsafe-swiftshader")
        chrome_options.add_argument("User-Agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        chrome_options.add_argument("Accept-Language=en-US,en;q=0.9")
        chrome_options.add_argument("Accept-Encoding=gzip, deflate, br")
        chrome_options.add_argument("Connection=keep-alive")
        chrome_options.add_argument("Upgrade-Insecure-Requests=1")
        chrome_options.add_argument("Dnt=1")
        chrome_options.add_argument("Accept=*/*")
        chrome_options.add_argument("Cache-Control=no-cache")
        chrome_options.add_argument("Pragma=no-cache")
        chrome_options.add_argument("TE=Trailers")
        return chrome_options

    def extract_exchange_rate_for_currency_pair(self, url, base_currency, target_currency):
        """
        Extracts the exchange rate for the specified currency pair from the given URL.
        """
        try:
            # Open the webpage
            self.driver.get(url)
            print(f"Opened URL: {url}")

            # Wait for a few seconds to ensure the page is loaded
            time.sleep(2)  # Wait for 3 seconds before extracting data (adjust as needed)

            # Wait for the body element to load
            WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

            # Locate the section containing the exchange rate data
            body_element = self.driver.find_element(By.TAG_NAME, "body")
            section_div = body_element.find_element(By.CLASS_NAME, "sectionx")
            month_year_element = section_div.find_element(By.XPATH, f".//h3[text()='{self.current_month_year}']")
            table_element = month_year_element.find_element(By.XPATH, "following-sibling::table[@id='hist']")

            # Try to find the row with yesterday's date
            row_element, target_date = self.get_exchange_rate_row(table_element)

            if row_element:
                # Extract the exchange rate from the row
                closing_rate_text = row_element.find_element(By.XPATH, ".//td[2]").text
                exchange_rate = self.extract_rate_from_text(closing_rate_text)

                if exchange_rate:
                    # Update the customer data with the exchange rate
                    self.update_currency_data(base_currency, target_currency, exchange_rate, target_date)
                    return (base_currency, target_currency, self.yesterday_date_str)

            return None

        except Exception as e:
            print(f"Error occurred for URL {url}: {e}")
            return None

    def get_exchange_rate_row(self, table_element):
        """
        Attempts to find the row with yesterday's date, retrying up to 7 times if not found.
        """
        attempts = 0
        target_date = self.yesterday_date_str

        
        try:
            row_element = table_element.find_element(By.XPATH, f".//tr[@id='{target_date}']")
            return row_element, target_date
        except Exception as e:
            print(f"Error occurred for {target_date}: {type(e).__name__}.")
           # target_date_obj = datetime.strptime(target_date, '%d-%m-%Y') - timedelta(days=1)
           # target_date = target_date_obj.strftime('%d-%m-%Y')
           # attempts += 1
        return None

    def extract_rate_from_text(self, text):
        """
        Extracts the exchange rate from the provided text using regex.
        """
        match = re.search(r"([^\d\s]+)\d\s([A-Za-z]{3})\s*=\s*([^\s]+)", text)
        if match:
            return match.group(3)
        return None

    def update_currency_data(self, base_currency, target_currency, exchange_rate, target_date):
        """
        Updates the currency data with the extracted exchange rate.
        """
        for column in self.Exchange_rate:
            if column['From currency'] == base_currency and column['To-currency'] == target_currency and column['To-currency'] == 'INR':
                column['Valid from'] = target_date.replace("-",'.')
                cleaned_value = re.sub(r"[^\d.]+", "", exchange_rate)
                #column['Direct Quote'] = re.sub(r"[^\d.]+", "", exchange_rate)
                column['Direct Quote'] = cleaned_value
                column['Indirect Quote'] = 0
                
            if column['From currency'] == target_currency and column['To-currency'] == base_currency and column['From currency'] == 'INR':
                #print("####")
                column['Valid from'] = target_date.replace("-",'.')
               # cleaned_value = re.sub(r"[^\d.]+", "", exchange_rate)
               # column['Indirect Quote'] = float(1 / float(cleaned_value))
                column['Indirect Quote'] = cleaned_value
                column['Direct Quote'] = 0
                

    def create_or_update_excel_file(self):
        """
        Creates a new Excel file or updates an existing one with the currency data.
        """
        for file_path in file_paths:
            # Check if the file exists, if not, create it
            if not os.path.exists(file_path):
                wb = Workbook()
                ws = wb.active
                ws.title = 'Exchange rate'

                # Define column headers
                headers = ['Exchange rate type', 'Valid from', 'Indirect Quote', 'From currency', 'Direct Quote', 'To-currency']
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num, value=header)

                wb.save(file_path)

            # Load the existing workbook and worksheet
            wb = openpyxl.load_workbook(file_path)
            ws = wb['Exchange rate']

            # Clear existing data except headers
            for row in range(2, ws.max_row + 1):  # Starting from row 2 to preserve the headers
                for col in range(1, 7):  # 6 columns for your data
                    ws.cell(row=row, column=col, value=None)

            # Add currency data to the worksheet
            if os.path.basename(file_path) == 'list_of_Exchange_rate.xlsx':
                index = 2
                for rate in self.Exchange_rate:
                    if rate['Valid from'] is not None:
                        
                        ws[f'A{index}'] = rate['Exchange rate type']
                        ws[f'B{index}'] = rate['Valid from']
                        ws[f'C{index}'] = rate['Indirect Quote']
                        ws[f'D{index}'] = rate['From currency']
                        ws[f'E{index}'] = rate['Direct Quote']
                        ws[f'F{index}'] = rate['To-currency']
                        
                        index = index + 1
                    else :
                        continue
                
            # else:
            #     index = ws.max_row+1
            #     for rate in self.Exchange_rate:
            #         if rate['Valid from'] is not None:
                        
            #             ws[f'A{index}'] = rate['Exchange rate type']
            #             ws[f'B{index}'] = rate['Valid from']
            #             ws[f'C{index}'] = rate['Indirect Quote']
            #             ws[f'D{index}'] = rate['From currency']
            #             ws[f'E{index}'] = rate['Direct Quote']
            #             ws[f'F{index}'] = rate['To-currency']
                        
            #             index = index + 1
            #         else :
            #             continue    

            # Format the 'Valid from' column as text
            valid_from_column = 2  # 'Valid from' is in the second column (B)
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=valid_from_column)
                cell.number_format = '@'  # Set format to text
                # if cell.value:
                #     cell.value = f"{cell.value:0>18}"  # Zero pad if needed

            # Save the workbook after all updates
            wb.save(file_path)
            print(f"Data successfully written to {file_path}")

    def send_email_with_attachment(self,email_config):
        # Create the email message object
        msg = MIMEMultipart()
        msg['From'] = email_config['from']
        msg['To'] = email_config["to"]
        msg['Subject'] = email_config["subject"]

        # Attach the body text to the email
        msg.attach(MIMEText(email_config['body'], 'plain'))

        # Attach the Excel file
        try:
            with open(email_config['file_path'], "rb") as file:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(file.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(email_config["file_path"])}')
                msg.attach(part)
        except FileNotFoundError:
            print(f"Error: The file at {email_config['file_path']} was not found.")
            return
        except Exception as e:
            print(f"Error attaching file: {e}")
            return

        # Set up the SMTP server
        try:
            server = smtplib.SMTP('smtp.gmail.com', email_config["smtp_port"])  # Change to your email provider's SMTP server
            server.starttls()  # Use TLS for security
            server.login(email_config['from'], email_config['password'])  # Log in to the SMTP server
            text = msg.as_string()  # Convert the message to string format
            server.sendmail(email_config['from'], email_config['to'], text)  # Send the email
            server.quit()  # Close the server connection
            print(f"Email sent successfully to {email_config['to']}")
        except smtplib.SMTPAuthenticationError:
            print("Error: Authentication failed. Please check your email and password.")
        except Exception as e:
            print(f"Failed to send email: {e}")



    def run(self,email_config):
        """
        Runs the scraper: extracts exchange rates for all currency pairs and updates the Excel file.
        """
        # Create or update the Excel file
        #self.create_or_update_excel_file()

        #Iterate through currency data and fetch the exchange rate for each currency pair
        for curr in self.Exchange_rate:
            
            base_currency = curr['From currency']
            target_currency = curr['To-currency']

            if base_currency == 'INR':
                break
            # Construct the URL dynamically based on currency pair
            url = f"https://www.exchangerates.org.uk/{base_currency}-{target_currency}-spot-exchange-rates-history-2024.html#{self.yesterday_date_str}"

            # Extract the exchange rate and update customer data
            self.extract_exchange_rate_for_currency_pair(url, base_currency, target_currency)

        # After processing, save the updated Excel file
        self.create_or_update_excel_file()

        # Close the WebDriver after processing all URLs
        self.driver.quit()
        print("WebDriver closed.")
        
        
       # self.send_email_with_attachment(email_config)
    


# Example usage
if __name__ == "__main__":
    
    file_paths = ['G:\\Shared drives\\Metal Prices\\Market_price\\list_of_Exchange_rate.xlsx','G:\\Shared drives\\Metal Prices\\Market_price\\list_of_Exchange_rate_history.xlsx']
    
    Exchange_rate= [
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "USD", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "EUR", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "GBP", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "AED", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "AUD", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "CAD", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "CHF", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "CNY", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "DKK", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "HKD", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "JPY", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "NOK", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "NZD", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "OMR", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "RUB", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "SAR", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "SEK", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "SGD", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "ZAR", "Direct Quote": None, "To-currency": "INR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "AED"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "AUD"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "CAD"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "CHF"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "CNY"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "DKK"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "EUR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "GBP"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "HKD"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "JPY"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "NOK"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "NZD"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "OMR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "RUB"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "SAR"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "SEK"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "SGD"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "USD"},
    {"Exchange rate type": "M", "Valid from": None, "Indirect Quote": None, "From currency": "INR", "Direct Quote": None, "To-currency": "ZAR"}
]

    email_config = {
        'subject' :"Test Email with Attachment",
        'body' : "This is the body of a mail",
        'from': "saket.verma@necoindia.com",
        'to': "saket.verma@necoindia.com",
       # 'bcc': "saket.verma@necoindia.com",
        'smtp_server': 'smtp.gmail.com',
        'smtp_port': 587,
        'password': "pxou ynvf ricv vvei" , 
        'file_path': f"{file_paths[0]}"
    }

    
    

    scraper = CurrencyExchangeScraper(Exchange_rate, file_paths)
    scraper.run(email_config)








