import openpyxl
from datetime import datetime, timedelta
from openpyxl.styles import Alignment
import os

# List of Titles for which we want to change the PublishedDate
titles_to_update = [
    "Premium Low Vol",
    "Low Vol HCC",
    "Low Vol PCI",
    "Mid Vol PCI",
    "Semi Soft",
    "Steam Coal - 6000 NAR"
]

# Function to update the date in the "PublishedDate" column to the previous date for specific titles
def update_dates_to_previous(file_path):
    # Open the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # Use the active sheet

    # Find the column index for "PublishedDate" and "Title"
    published_date_col = None
    title_col = None
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header == 'PublishedDate':
            published_date_col = col
        elif header == 'Title':
            title_col = col

    if published_date_col is None:
        raise ValueError("The column 'PublishedDate' was not found.")
    if title_col is None:
        raise ValueError("The column 'Title' was not found.")

    # Iterate through the rows (starting from row 2 to skip header)
    for row in range(2, sheet.max_row + 1):
        title_cell = sheet.cell(row=row, column=title_col)
        published_date_cell = sheet.cell(row=row, column=published_date_col)
        
        title_value = title_cell.value
        date_value = str(published_date_cell.value)  # Make sure it's a string

        # Check if the title matches one of the specified titles
        if title_value in titles_to_update:
            try:
                # Ensure the date is in the correct format (YYYYMMDD)
                if len(date_value) == 8:  # Check for valid date length
                    # Convert to a datetime object
                    current_date = datetime.strptime(date_value, '%Y%m%d')
                    
                    # Subtract one day to get the previous date
                    previous_date = current_date - timedelta(days=1)
                    
                    # Convert the previous date back to YYYYMMDD format and update the cell
                    previous_date_str = previous_date.strftime('%Y%m%d')
                    published_date_cell.value = previous_date_str
                    published_date_cell.number_format = '@'
                    
                    # Set the alignment to right
                    published_date_cell.alignment = Alignment(horizontal='right', vertical='center')
            except Exception as e:
                print(f"Error processing row {row}: {e}")

    # Save the updated file to a new file (or overwrite the existing one)
    workbook.save("G:\\Shared drives\\Metal Prices\\Market_price\\updated_" + os.path.basename(file_path))

# Example usage
file_path = "G:\\Shared drives\\Metal Prices\\Market_price\\Combined_history.xlsx"  # Replace with your file path
update_dates_to_previous(file_path)
