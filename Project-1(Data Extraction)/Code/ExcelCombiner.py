import pandas as pd
import os
import openpyxl
from datetime import datetime , timedelta
import time

class ExcelCombiner:
    def __init__(self, directory, file_names, common_columns,load_directory,date):
        
        """
        Initialize the ExcelCombiner object with the specified parameters.
        
        Args:
        - directory (str): The directory where the Excel files are stored.
        - file_names (list): List of file names to be processed.
        - common_columns (list): List of columns to be included when combining files.
        - load_directory (str): The secondary directory to load files if not found in the primary directory.
        - date (str): The current date (in the format YYYYMMDD), which will be appended to the output files' names.
        """
        self.directory = directory
        self.file_names = file_names
        self.combined_df = None
        self.common_columns = common_columns
        self.load_directory = load_directory
        self.date = date
        
    def update_dates_in_xlsx(self, directory):
        
        """
        Update or create a 'PublishedDate' column in the specified Excel files within the directory.
        The 'PublishedDate' will be set to yesterday's date for 'Platts_data.xlsx' and 'Argus_data.xlsx',
        and today's date for the other files. It also formats the 'ID' column as text.

        Args:
        - directory (str): The directory containing the Excel files.
        """
        
        # Look for the filtered_data.csv file in the given directory
        print(self.file_names)
        for file in os.listdir(directory):
            base_name = os.path.basename(file)
            if base_name in self.file_names:
                file_path = os.path.join(directory, base_name)
                print(file_path)

                # Check if the file exists
                if not os.path.isfile(file_path):
                    print(f"\nFile '{file_path}' not found in directory '{directory}'.")
                    continue  # Use continue instead of return to skip to the next file

                # Load the Excel file
                df = pd.read_excel(file_path)
                #current_date = datetime.now().strftime('%Y%m%d')
                current_date = datetime.now()
                
                if base_name == 'Platts_data.xlsx' or base_name == 'Argus_data.xlsx':
                    yesterday_date = current_date - timedelta(days=1)
                    
                    yesterday_date = yesterday_date.strftime('%Y%m%d')
                    if 'PublishedDate' in df.columns:
                        # Update the 'PublishedDate' column for all rows
                        df['PublishedDate'] = yesterday_date
                        print("\nPublishedDate updated successfully.")
                    else:
                        df.insert(0, 'PublishedDate', yesterday_date)
                        print("\nColumn 'PublishedDate' not found in the file, so created it.")
                        
                else:
                    current_date = current_date.strftime('%Y%m%d')
                    if 'PublishedDate' in df.columns:
                        # Update the 'PublishedDate' column for all rows
                        df['PublishedDate'] = current_date
                        print("\nPublishedDate updated successfully.")
                    else:
                        df.insert(0, 'PublishedDate', current_date)
                        print("\nColumn 'PublishedDate' not found in the file, so created it.")

                # Save the updated DataFrame back to the Excel file
                df.to_excel(file_path, index=False)

                wb = openpyxl.load_workbook(file_path)
                ws = wb.active

                # Specify the columns you want to format as text
                text_columns = ["ID"]

                # Apply text formatting to specified columns, including the header
                for col in text_columns:
                    if col in df.columns:
                        col_index = df.columns.get_loc(col) + 1  # Get column index (1-based)

                        # Set the header to text format
                        header_cell = ws.cell(row=1, column=col_index)  # Assuming headers are in the first row
                        header_cell.number_format = '@'  # Set format to Text

                        # Set the data cells to text format
                        for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):  # Skip header
                            for cell in row:
                                cell.number_format = '@'  # Set format to Text
                                cell.value = f"{cell.value:0>18}"

                # Save the updated workbook
                wb.save(file_path)
                print(f"Data successfully written to {file_path}")
    
    def update_date_BIGmint(self):
        
        """
        Update or create a 'PublishedDate' column in the specified BIGmint file.
        The 'PublishedDate' will be set to today's date. This function is specific to the last file in the list.
        The method also handles file loading, updating, and saving.

        """
        name = self.file_names[-1]
        file = os.path.join(self.load_directory, name)
        print(f"File path: {file}")
    
        try:
            # Load the Excel file
            df = pd.read_excel(file)
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return
        
        current_date = datetime.now().strftime('%Y%m%d')
        #current_date = '20241024'  # Example date

        # Check if 'PublishedDate' exists, and update or insert as necessary
        if 'PublishedDate' in df.columns:
            df['PublishedDate'] = current_date
            print("PublishedDate updated successfully.")
        else:
            df.insert(0, 'PublishedDate', current_date)
            print("Column 'PublishedDate' not found, created it.")

        try:
            # Save the updated DataFrame back to the Excel file
            df.to_excel(file, index=False)
            print("DataFrame saved to Excel.")
        except Exception as e:
            print(f"Error saving DataFrame to Excel: {e}")
            return
            

    def combine_files(self):
        """
        Combine multiple Excel files based on common columns into a single DataFrame.
        It reads each Excel file from the specified directories and appends data
        that contains the common columns. This allows merging of files with a similar structure.
        
        """
        dataframes = []
    
        for file_name in self.file_names:
            file_path = os.path.join(self.directory, file_name)
            df = None  # Initialize df to None

            try:
                # Attempt to read the Excel file from the main directory
                df = pd.read_excel(file_path)
            except FileNotFoundError:
                # If not found, attempt to read from the load directory
                file_path = os.path.join(self.load_directory, file_name)
                try:
                    df = pd.read_excel(file_path)
                except FileNotFoundError:
                    print(f"File not found: {file_name} in both directories.")
                except ValueError as ve:
                    print(f"Error reading {file_name}: {ve}")

            if df is not None:
                # Ensure the common columns exist in the DataFrame
                common_columns_exist = [col for col in self.common_columns if col in df.columns]
                if common_columns_exist:
                    filtered_df = df[common_columns_exist]
                    dataframes.append(filtered_df)
                else:
                    print(f"Common columns not found in {file_name}: {self.common_columns}")

        if dataframes:
            self.combined_df = pd.concat(dataframes, ignore_index=True)
            print("\nFiles combined successfully based on the specified pattern.")
        else:
            print("No valid dataframes to combine.")
            
        

    def save_combined_file(self, output_file):
        """
        Save the combined DataFrame into an Excel file, both with and without the current date appended to the file name.
        The method also formats the 'ID' column as text and applies padding to ensure values are 18 digits long.
        
        Args:
        - output_file (str): The name for the output file (without extension).
        """
        if self.combined_df is not None:
            #print(self.combined_df)
            file_names = []  # Initialize a list to store file names

            # Save without date
            file_name = os.path.join(self.directory, f'{output_file}.xlsx')
            self.combined_df.to_excel(file_name, index=False)
            file_names.append(file_name)
            print(f"\nCombined data saved as '{file_name}'.")

            # Save with date using the same variable
            file_name = os.path.join(self.directory, f'{output_file}_{self.date}.xlsx')
            self.combined_df.to_excel(file_name, index=False)
            file_names.append(file_name)
            print(f"\nCombined data saved as '{file_name}'.")

            # Format the files
            for file in file_names:
                try:
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active

                    # Specify the columns you want to format as text
                    text_columns = ["ID"]

                    # Apply text formatting to specified columns, including the header
                    for col in text_columns:
                        if col in self.combined_df.columns:
                            col_index = self.combined_df.columns.get_loc(col) + 1  # Get column index (1-based)

                            # Set the header to text format
                            header_cell = ws.cell(row=1, column=col_index)  # Assuming headers are in the first row
                            header_cell.number_format = '@'  # Set format to Text

                            # Set the data cells to text format
                            for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):  # Skip header
                                for cell in row:
                                    cell.number_format = '@'  # Set format to Text
                                    cell.value = f"{cell.value:0>18}"

                    # Save the updated workbook
                    wb.save(file)
                    print(f"Data successfully written to {file}")

                except Exception as e:
                    print(f"An error occurred while processing {file}: {e}")

        else:
            print("No data to save. Please run combine_files() first.")


    def delete_original_files(self):
        """
        Delete the original Excel files after processing and saving the combined data.
        It first attempts to delete the files from the primary directory, and if not found,
        it tries the secondary directory (load_directory).
        """
        for file_name in self.file_names:
            file_path = os.path.join(self.directory, file_name)
            print(f"Attempting to remove: {file_path}")

            # Try to delete the file, with retries for locked files
            for attempt in range(3):  # Retry up to 3 times
                try:
                    os.remove(file_path)
                    print(f"Successfully removed: {file_path}")
                    break  # Exit loop if successful
                except FileNotFoundError:
                    print(f"File not found: {file_path}. Trying a different path...")
                    # Try deleting from the alternative directory
                    alternative_path = os.path.join(self.load_directory, file_name)
                    try:
                        os.remove(alternative_path)
                        print(f"Successfully removed: {alternative_path}")
                        break  # Exit loop if successful
                    except FileNotFoundError:
                        print(f"Also not found at alternative path: {alternative_path}")
                        print(f"File '{file_name}' was not found in either directory.")
                        break  # Exit loop if not found in either directory
                except PermissionError:
                    print(f"PermissionError: The file is in use or locked. Retrying in 1 second...")
                    time.sleep(1)  # Wait before retrying
                except Exception as e:
                    print(f"Unexpected error occurred while deleting {file_name}: {e}")
                    break  # Exit the retry loop on unexpected errors


# Example usage
# if __name__ == "__main__":
#     directory = r"D:\USER PROFILE DATA\Desktop\Project-1\Data\\"
#     file_names = [
#         'Argus_data.xlsx',
#         'LME_data.xlsx',
#         'Nalco_data.xlsx',
#         'Platts_data.xlsx'
#     ]
#     common_columns = ['Column1', 'Column2', 'Column3']  # List of columns to check
#     pattern = 'Common_Pattern'  # Pattern to look for

#     combiner = ExcelCombiner(directory, file_names, common_columns, pattern)
#     combiner.combine_files()
#     combiner.save_combined_file('Combined_data.xlsx')
#     combiner.delete_original_files()
